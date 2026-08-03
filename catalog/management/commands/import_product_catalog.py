from hashlib import md5
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from catalog.models import AttributeDataType, Category, CategoryAttribute


COMMON_ATTRIBUTES = {'SKU', '品牌', '产品名称', '产品型号', '产品规格', '销售价格'}

ATTRIBUTE_CODE_MAP = {
    'SKU': 'sku',
    '品牌': 'brand',
    '产品名称': 'product_name',
    '产品型号': 'model',
    '产品规格': 'spec',
    '销售价格': 'price',
    '颜色': 'color',
    '包材': 'package_material',
    '克重': 'gram_weight',
    '箱规': 'carton_spec',
    '碘值': 'iodine_value',
    '尺寸': 'size',
    '孔径': 'pore_size',
    '亚兰值': 'methylene_blue_value',
}

NUMERIC_ATTRIBUTES = {'克重', '碘值', '亚兰值'}


def clean_text(value):
    if value is None:
        return ''
    return str(value).strip()


def stable_slug(*parts):
    label = '-'.join(part for part in parts if part)
    ascii_slug = slugify(label)
    if ascii_slug:
        return ascii_slug[:140]
    digest = md5(label.encode('utf-8')).hexdigest()[:12]
    return f'cat-{digest}'


def attribute_code(name):
    mapped = ATTRIBUTE_CODE_MAP.get(name)
    if mapped:
        return mapped
    ascii_code = slugify(name).replace('-', '_')
    if ascii_code:
        return ascii_code[:100]
    digest = md5(name.encode('utf-8')).hexdigest()[:12]
    return f'attr_{digest}'


def attribute_type(name):
    if name == '销售价格':
        return AttributeDataType.PRICE
    if name in NUMERIC_ATTRIBUTES:
        return AttributeDataType.NUMBER
    return AttributeDataType.TEXT


class Command(BaseCommand):
    help = '从产品目录 Excel 导入三级分类和类目属性模板'

    def add_arguments(self, parser):
        parser.add_argument('file', help='产品目录 Excel 文件路径')
        parser.add_argument('--dry-run', action='store_true', help='只校验不写入数据库')

    def handle(self, *args, **options):
        file_path = Path(options['file'])
        if not file_path.exists():
            self.stderr.write(self.style.ERROR(f'文件不存在：{file_path}'))
            return

        wb = load_workbook(file_path, data_only=True)
        if '总目录' not in wb.sheetnames:
            self.stderr.write(self.style.ERROR('缺少工作表：总目录'))
            return

        created_categories = 0
        updated_categories = 0
        created_attributes = 0
        updated_attributes = 0
        failures = []

        with transaction.atomic():
            category_by_name = {}
            ws = wb['总目录']
            for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                level1, level2, level3 = [clean_text(value) for value in row[:3]]
                if not any((level1, level2, level3)):
                    continue
                if not level1:
                    failures.append((row_no, '一级目录不能为空'))
                    continue

                parent = None
                path = []
                for depth, name in enumerate((level1, level2, level3), start=1):
                    if not name:
                        continue
                    path.append(name)
                    slug = stable_slug(*path)
                    category, created = Category.objects.update_or_create(
                        slug=slug,
                        defaults={
                            'name': name,
                            'parent': parent,
                            'sort_order': row_no * 10 + depth,
                            'is_active': True,
                        },
                    )
                    category_by_name.setdefault(name, category)
                    category_by_name['/'.join(path)] = category
                    created_categories += int(created)
                    updated_categories += int(not created)
                    parent = category

            for sheet_name in wb.sheetnames:
                if sheet_name == '总目录':
                    continue
                target_level = None
                category_name = sheet_name
                if sheet_name.startswith('一级-'):
                    target_level = 1
                    category_name = sheet_name.removeprefix('一级-')
                elif sheet_name.startswith('二级-'):
                    target_level = 2
                    category_name = sheet_name.removeprefix('二级-')
                elif sheet_name.startswith('三级-'):
                    target_level = 3
                    category_name = sheet_name.removeprefix('三级-')

                candidates = Category.objects.filter(name=category_name, is_active=True)
                if target_level:
                    candidates = [category for category in candidates if category.level == target_level]
                else:
                    candidates = list(candidates)
                category = candidates[0] if candidates else category_by_name.get(category_name)
                if not category and target_level == 2 and category_name.endswith('其他'):
                    parent_name = category_name.removesuffix('其他')
                    parent = Category.objects.filter(name=parent_name, parent__isnull=True, is_active=True).first()
                    if parent:
                        category, created = Category.objects.update_or_create(
                            slug=stable_slug(parent.name, category_name),
                            defaults={
                                'name': category_name,
                                'parent': parent,
                                'sort_order': 9990,
                                'is_active': True,
                            },
                        )
                        created_categories += int(created)
                        updated_categories += int(not created)
                if not category:
                    failures.append((sheet_name, f'找不到分类：{category_name}'))
                    continue

                headers = [clean_text(value) for value in next(wb[sheet_name].iter_rows(min_row=1, max_row=1, values_only=True))]
                special_headers = [header for header in headers if header and header not in COMMON_ATTRIBUTES]
                for index, header in enumerate(special_headers, start=1):
                    _, created = CategoryAttribute.objects.update_or_create(
                        category=category,
                        code=attribute_code(header),
                        defaults={
                            'name': header,
                            'data_type': attribute_type(header),
                            'is_required': False,
                            'is_filterable': True,
                            'is_list_visible': True,
                            'is_detail_visible': True,
                            'sort_order': index * 10,
                            'is_active': True,
                        },
                    )
                    created_attributes += int(created)
                    updated_attributes += int(not created)

            if options['dry_run']:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS('导入完成' if not options['dry_run'] else '校验完成，未写入数据库'))
        self.stdout.write(f'分类新增 {created_categories}，更新 {updated_categories}')
        self.stdout.write(f'属性新增 {created_attributes}，更新 {updated_attributes}')
        self.stdout.write(f'失败 {len(failures)}')
        for row_no, reason in failures[:50]:
            self.stdout.write(f'{row_no}：{reason}')
        if len(failures) > 50:
            self.stdout.write(f'还有 {len(failures) - 50} 条失败未显示')
