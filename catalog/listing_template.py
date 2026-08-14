"""上架模板（14 列固定）专用导入器。

模板列顺序固定为：
    一级目录 | 二级目录 | 三级目录 | 商品编码 | 品牌 | 产品名称
    | 制造商型号 | 容量 | 颜色 | 含税价 | 销售单位 | 包规 | MOQ | MPQ

去重规则：
    - Product 视为"同一类型系列"当且仅当
      (品牌, 产品名称, 制造商型号) 三个字段完全相同
      —— 这是按用户业务定义的"4 字段相同 → 同类型"，但色号作为 SKU 区分。
    - 不允许跨 Product 合并：相同三字段只允许有一个 Product，
      其下挂多个 SKU（不同颜色 / 不同容量）。
    - Product.style_code 由 `f"<品牌>-<产品名>-<型号>"` 拼接生成，
      自动加数字后缀避免碰撞。
    - SKU 唯一键 = `商品编码`，与 SKU.internal_sku_code 一致。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from .models import Category, Product, PublishStatus, SKU, StockStatus


EXPECTED_HEADERS = (
    '一级目录',
    '二级目录',
    '三级目录',
    '商品编码',
    '品牌',
    '产品名称',
    '制造商型号',
    '容量',
    '颜色',
    '含税价',
    '销售单位',
    '包规',
    'MOQ',
    'MPQ',
)

_INVALID_FN_CHARS = re.compile(r'[\\/:*?"<>|]+')
_SPACES = re.compile(r'\s+')
_SAFE_NULL = {'/', '—', '-', '无', '无 /', ' /'}


@dataclass
class ListingImportResult:
    created_products: int = 0
    updated_products: int = 0
    created_skus: int = 0
    updated_skus: int = 0
    created_categories: int = 0
    failures: list[str] = field(default_factory=list)


def clean_text(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    if text in _SAFE_NULL:
        return ''
    return text


def to_decimal(value) -> Decimal | None:
    text = clean_text(value)
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return None


def ensure_safe_name(value: str, fallback: str) -> str:
    cleaned = _INVALID_FN_CHARS.sub('-', value or '').strip()
    cleaned = _SPACES.sub(' ', cleaned)
    return cleaned or fallback


def category_chain_exists(category: Category | None, names: list[str]) -> bool:
    """检查 product.category 是否位于 names 链的某一级祖先下。"""
    if not category or not names:
        return False
    needle = [n for n in names if n]
    if not needle:
        return False
    current = category
    for ancestor_name in reversed(needle):
        if not current or current.name != ancestor_name:
            return False
        current = current.parent
    return True


def get_or_create_category_chain(l1: str, l2: str, l3: str) -> Category:
    """按一级/二级/三级 找/创建分类链路。

    命中规则：
    1. 完全匹配三级 → 用
    2. 找不到三级则尝试匹配二级作为叶子（用于补二级分类商品）
    3. 都没有 → 按层级自顶向下创建
    """
    l1 = clean_text(l1)
    l2 = clean_text(l2)
    l3 = clean_text(l3)

    if l3:
        candidates = Category.objects.filter(name=l3, is_active=True)
        for cat in candidates.select_related('parent', 'parent__parent'):
            if cat.parent and cat.parent.name == l2 and cat.parent.parent and cat.parent.parent.name == l1:
                return cat
        parent_l2 = _ensure_category(l1, l2)
        return _ensure_child(parent_l2, l3)
    if l2:
        candidates = Category.objects.filter(name=l2, is_active=True)
        for cat in candidates.select_related('parent'):
            if cat.parent and cat.parent.name == l1:
                return cat
        return _ensure_category(l1, l2)
    if l1:
        candidates = Category.objects.filter(name=l1, parent__isnull=True, is_active=True)
        if candidates.exists():
            return candidates.first()
        return _ensure_root(l1)

    fallback, _ = Category.objects.get_or_create(slug='uncategorized', defaults={'name': '未分类'})
    return fallback


def _ensure_root(name: str) -> Category:
    existing = Category.objects.filter(parent__isnull=True, is_active=True).filter(name=name).first()
    if existing:
        return existing
    slug = _unique_slug(name)
    return Category.objects.create(name=name, slug=slug, parent=None)


def _ensure_category(l1_name: str, l2_name: str) -> Category:
    parent = _ensure_root(l1_name)
    existing = Category.objects.filter(parent=parent, name=l2_name, is_active=True).first()
    if existing:
        return existing
    return Category.objects.create(name=l2_name, slug=_unique_slug(f'{l1_name}-{l2_name}'), parent=parent)


def _ensure_child(parent: Category, name: str) -> Category:
    existing = Category.objects.filter(parent=parent, name=name, is_active=True).first()
    if existing:
        return existing
    return Category.objects.create(name=name, slug=_unique_slug(f'{parent.name}-{name}'), parent=parent)


def _unique_slug(name: str) -> str:
    base = slugify(name, allow_unicode=True) or 'item'
    slug = base
    index = 2
    while Category.objects.filter(slug=slug).exists():
        slug = f'{base}-{index}'
        index += 1
    return slug


def build_product_style_code(brand: str, product_name: str, model: str) -> str:
    parts = [ensure_safe_name(brand, 'X'), ensure_safe_name(product_name, 'X'), ensure_safe_name(str(model or ''), 'X')]
    raw = '-'.join(p for p in parts if p)
    raw = _SPACES.sub('-', raw)
    base = raw[:80]
    candidate = base
    index = 2
    while Product.objects.filter(style_code=candidate).exists():
        suffix = f'-{index}'
        if len(base) + len(suffix) > 80:
            base = base[: 80 - len(suffix)]
        candidate = f'{base}{suffix}'
        index += 1
    return candidate


def find_product_by_signature(brand: str, product_name: str, model: str) -> Product | None:
    qs = Product.objects.filter(
        brand__iexact=brand,
        name__iexact=product_name,
    )
    qs = [p for p in qs if (p.manufacturer_model or '').strip() == model.strip()]
    return qs[0] if qs else None


def import_listing_workbook(
    uploaded_file,
    *,
    dry_run: bool = False,
    source_file_name: str = '',
) -> ListingImportResult:
    result = ListingImportResult()
    if hasattr(uploaded_file, 'read'):
        data = uploaded_file.read()
        workbook = load_workbook(BytesIO(data), data_only=True)
    else:
        workbook = load_workbook(uploaded_file, data_only=True)
    if not workbook.sheetnames:
        result.failures.append('Excel 文件中没有工作表。')
        return result

    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        result.failures.append('工作表为空。')
        return result

    headers = [clean_text(cell) for cell in rows[0]]
    if headers != list(EXPECTED_HEADERS):
        result.failures.append(
            '表头不正确。期望：' + ' / '.join(EXPECTED_HEADERS)
        )
        return result

    seen_sku_codes: set[str] = set()
    product_cache: dict[tuple[str, str, str], Product] = {}

    with transaction.atomic():
        for row_index, row in enumerate(rows[1:], start=2):
            if row is None or all(cell in (None, '') for cell in row):
                continue
            values = list(row)
            # Pad to expected length to avoid IndexError
            while len(values) < len(EXPECTED_HEADERS):
                values.append('')

            (
                level1, level2, level3, sku_code, brand, product_name,
                manufacturer_model, capacity, color, price_raw, unit,
                package_spec, moq_raw, mpq_raw,
            ) = values

            sku_code = clean_text(sku_code)
            brand = clean_text(brand)
            product_name = clean_text(product_name)
            manufacturer_model = clean_text(manufacturer_model)
            capacity = clean_text(capacity)
            color = clean_text(color)
            unit = clean_text(unit)
            package_spec = clean_text(package_spec)

            if not sku_code:
                result.failures.append(f'第 {row_index} 行：商品编码不能为空。')
                continue
            if sku_code in seen_sku_codes:
                result.failures.append(f'第 {row_index} 行：商品编码 "{sku_code}" 在表中重复。')
                continue
            if not brand or not product_name or not manufacturer_model:
                result.failures.append(
                    f'第 {row_index} 行：品牌、产品名称、制造商型号均不能为空。'
                )
                continue

            seen_sku_codes.add(sku_code)
            signature = (brand.casefold(), product_name.casefold(), manufacturer_model.casefold())
            category = None
            try:
                category = get_or_create_category_chain(level1, level2, level3)
            except Exception as exc:  # pragma: no cover - defensive logging
                result.failures.append(f'第 {row_index} 行：分类处理失败：{exc}')
                continue
            if category and not category.is_active:
                category.is_active = True
                category.save(update_fields=['is_active'])

            product = product_cache.get(signature)
            if product is None:
                product = find_product_by_signature(brand, product_name, manufacturer_model)
                if product is not None:
                    product_cache[signature] = product

            price = to_decimal(price_raw)
            if price is None:
                result.failures.append(
                    f'第 {row_index} 行：含税价 "{price_raw}" 不是有效金额。'
                )
                continue

            moq_value = to_decimal(moq_raw) or Decimal('1')
            mpq_value = to_decimal(mpq_raw)

            if product is None:
                product = Product.objects.create(
                    name=product_name,
                    alias='',
                    style_code=build_product_style_code(brand, product_name, manufacturer_model),
                    cas_no='',
                    category=category,
                    brand=brand,
                    manufacturer_model=manufacturer_model,
                    capacity=capacity,
                    description='',
                    spec_summary='',
                    source_file_name=source_file_name,
                    status=PublishStatus.PUBLISHED,
                )
                product_cache[signature] = product
                result.created_products += 1
            else:
                updated_fields = []
                if product.name != product_name:
                    product.name = product_name
                    updated_fields.append('name')
                if product.brand != brand:
                    product.brand = brand
                    updated_fields.append('brand')
                if (product.manufacturer_model or '') != manufacturer_model:
                    product.manufacturer_model = manufacturer_model
                    updated_fields.append('manufacturer_model')
                if (product.capacity or '') != capacity and capacity:
                    product.capacity = capacity
                    updated_fields.append('capacity')
                if category and product.category_id != category.pk:
                    product.category = category
                    updated_fields.append('category')
                if updated_fields:
                    updated_fields.append('updated_at')
                    product.save(update_fields=updated_fields)
                    result.updated_products += 1

            # SKU 创建或更新
            existing_sku = SKU.objects.filter(internal_sku_code=sku_code).first()
            if existing_sku:
                updated = False
                sku_updates = []
                if existing_sku.product_id != product.pk:
                    existing_sku.product = product
                    sku_updates.append('product')
                    updated = True
                if (existing_sku.color or '') != color:
                    existing_sku.color = color
                    sku_updates.append('color')
                    updated = True
                if (existing_sku.capacity or '') != capacity:
                    existing_sku.capacity = capacity
                    sku_updates.append('capacity')
                    updated = True
                if (existing_sku.unit or '') != unit:
                    existing_sku.unit = unit
                    sku_updates.append('unit')
                    updated = True
                if (existing_sku.package_spec or '') != package_spec:
                    existing_sku.package_spec = package_spec
                    sku_updates.append('package_spec')
                    updated = True
                if existing_sku.price != price:
                    existing_sku.price = price
                    sku_updates.append('price')
                    updated = True
                if existing_sku.moq != moq_value:
                    existing_sku.moq = moq_value
                    sku_updates.append('moq')
                    updated = True
                if (existing_sku.mpq or None) != mpq_value:
                    existing_sku.mpq = mpq_value
                    sku_updates.append('mpq')
                    updated = True
                existing_sku.sku_attribute_text = color or existing_sku.sku_attribute_text
                existing_sku.source_goods_code = sku_code
                existing_sku.source_goods_name = product_name
                existing_sku.source_raw_row = {
                    **{header: clean_text(value) for header, value in zip(EXPECTED_HEADERS, values)},
                    '__row__': row_index,
                }
                sku_updates.extend(['sku_attribute_text', 'source_goods_code', 'source_goods_name', 'source_raw_row'])
                if updated or True:
                    sku_updates.append('updated_at')
                    existing_sku.save(update_fields=list(set(sku_updates)))
                result.updated_skus += 1
            else:
                SKU.objects.create(
                    product=product,
                    internal_sku_code=sku_code,
                    jst_sku_id=sku_code,
                    shop_sku_id=sku_code,
                    source_goods_code=sku_code,
                    source_style_code=product.style_code,
                    source_goods_name=product_name,
                    sku_attribute_text=color or '',
                    color=color,
                    capacity=capacity,
                    package_spec=package_spec,
                    unit=unit,
                    price=price,
                    moq=moq_value,
                    mpq=mpq_value,
                    order_step=mpq_value or Decimal('1'),
                    stock_status=StockStatus.CONFIRM,
                    inventory_sync_enabled=True,
                    source_raw_row={
                        **{header: clean_text(value) for header, value in zip(EXPECTED_HEADERS, values)},
                        '__row__': row_index,
                    },
                    status=PublishStatus.PUBLISHED,
                )
                result.created_skus += 1

        if dry_run:
            transaction.set_rollback(True)

    return result
