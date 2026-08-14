from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.contrib.staticfiles import finders
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook

from .models import Category, CategoryAttribute, Product, PublishStatus, SKU


class HomeCategoryIconTests(TestCase):
    def test_demo_second_level_categories_have_local_icons(self):
        for slug in ('spray-paint-demo', 'general-adhesive-demo'):
            with self.subTest(slug=slug):
                self.assertIsNotNone(finders.find(f'img/category-icons/{slug}.jpg'))


class ImportProductsCommandTests(TestCase):
    def test_imports_sample_shape_in_dev_mode(self):
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / 'products.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            ws.append(['款式编码', '商品编码', '商品名称', '商品简称', '颜色及规格', '颜色', '规格', '基本售价', '品牌', '分类', '商品状态', '库存同步'])
            ws.append(['STYLE-1', 'SKU-1', '测试密封胶 300ml', '密封胶', '白色 300ml', '白色', '300ml', None, '测试品牌', '', '启用', '开启'])
            wb.save(path)

            call_command('import_products', str(path), '--dev-allow-missing-price-category', verbosity=0)

        self.assertEqual(Product.objects.count(), 1)
        self.assertEqual(SKU.objects.count(), 1)
        sku = SKU.objects.get()
        self.assertEqual(sku.internal_sku_code, 'SKU-1')
        self.assertEqual(sku.jst_sku_id, 'SKU-1')

    def test_imports_only_valid_image_urls(self):
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / 'products.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            ws.append(['图片', '款式编码', '商品编码', '商品名称', '品牌'])
            ws.append(['https://example.com/product.jpg', 'STYLE-1', 'SKU-1', '测试产品', '测试品牌'])
            ws.append(['不是图片 URL', 'STYLE-2', 'SKU-2', '测试产品 2', '测试品牌'])
            wb.save(path)

            call_command('import_products', str(path), '--dev-allow-missing-price-category', verbosity=0)

        self.assertEqual(Product.objects.get(style_code='STYLE-1').remote_image_url, 'https://example.com/product.jpg')
        self.assertEqual(Product.objects.get(style_code='STYLE-2').remote_image_url, '')

    def test_imports_formal_product_attributes_from_category_template(self):
        category = Category.objects.create(name='活性炭', slug='carbon')
        CategoryAttribute.objects.create(category=category, name='碘值', code='iodine_value')
        CategoryAttribute.objects.create(category=category, name='孔径', code='pore_size')

        with TemporaryDirectory() as tmp:
            path = Path(tmp) / 'formal-products.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet1'
            ws.append(['SKU', '品牌', '产品名称', '产品型号', '产品规格', '销售价格', '分类', '碘值', '孔径'])
            ws.append(['SKU-C-1', '测试品牌', '柱状活性炭', 'AC-1', '25kg/袋', '128.5', '活性炭', '900', '4mm'])
            wb.save(path)

            call_command('import_products', str(path), verbosity=0)

        sku = SKU.objects.get(internal_sku_code='SKU-C-1')
        self.assertEqual(sku.price, Decimal('128.50'))
        self.assertEqual(sku.attributes, {'iodine_value': '900', 'pore_size': '4mm'})


class ImportProductCatalogCommandTests(TestCase):
    def test_imports_category_tree_and_attribute_templates(self):
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / 'catalog.xlsx'
            wb = Workbook()
            total = wb.active
            total.title = '总目录'
            total.append(['一级目录', '二级目录', '三级目录'])
            total.append(['车间化学品', '活性炭', '柱状活性炭'])
            active_carbon = wb.create_sheet('二级-活性炭')
            active_carbon.append(['SKU', '品牌', '产品名称', '碘值', '尺寸', '孔径', '亚兰值', '产品规格'])
            wb.save(path)

            call_command('import_product_catalog', str(path), verbosity=0)

        self.assertTrue(Category.objects.filter(name='车间化学品', parent__isnull=True).exists())
        active_carbon = Category.objects.get(name='活性炭')
        self.assertEqual(active_carbon.parent.name, '车间化学品')
        self.assertTrue(Category.objects.filter(name='柱状活性炭', parent=active_carbon).exists())
        self.assertEqual(
            set(active_carbon.attributes.values_list('code', flat=True)),
            {'iodine_value', 'size', 'pore_size', 'methylene_blue_value'},
        )


class ProductListViewTests(TestCase):
    def setUp(self):
        self.root = Category.objects.create(name='车间化学品', slug='workshop')
        self.mid = Category.objects.create(name='活性炭', slug='carbon', parent=self.root)
        self.leaf = Category.objects.create(name='柱状活性炭', slug='carbon-column', parent=self.mid)
        self.product = Product.objects.create(
            name='柱状活性炭 25kg',
            style_code='STYLE-C-1',
            brand='测试品牌',
            category=self.leaf,
            status=PublishStatus.PUBLISHED,
        )
        SKU.objects.create(
            product=self.product,
            internal_sku_code='SKU-C-1',
            jst_sku_id='SKU-C-1',
            status=PublishStatus.PUBLISHED,
        )

    def test_shows_subcategories_when_selected_category_has_children(self):
        response = self.client.get(reverse('catalog:product_list'), {'category': self.root.slug})
        self.assertContains(response, 'category-panel-label">分类</span>')
        self.assertContains(response, f'category={self.mid.slug}')
        self.assertNotContains(response, 'category-panel-label">品牌</span>')

    def test_shows_brands_when_selected_category_is_leaf(self):
        response = self.client.get(reverse('catalog:product_list'), {'category': self.leaf.slug})
        self.assertContains(response, 'category-panel-label">品牌</span>')
        self.assertContains(response, '测试品牌')
        self.assertNotContains(response, 'category-panel-label">分类</span>')

    def test_search_limits_category_panel_to_matching_products(self):
        other_leaf = Category.objects.create(name='其他分类', slug='other-leaf', parent=self.mid)
        Product.objects.create(
            name='其他产品',
            style_code='STYLE-C-2',
            brand='其他品牌',
            category=other_leaf,
            status=PublishStatus.PUBLISHED,
        )
        response = self.client.get(
            reverse('catalog:product_list'),
            {'category': self.root.slug, 'q': '柱状活性炭'},
        )
        chip_prefix = 'class="chip" href="?q=%E6%9F%B1%E7%8A%B6%E6%B4%BB%E6%80%A7%E7%82%AD&category='
        self.assertContains(response, f'{chip_prefix}{self.mid.slug}"')
        self.assertNotContains(response, f'{chip_prefix}{other_leaf.slug}"')

    def test_global_header_contains_hover_category_entry(self):
        response = self.client.get(reverse('catalog:product_list'))

        self.assertContains(response, 'class="header-search"')
        self.assertContains(response, 'class="cat-flyout-nav-head"')
        self.assertContains(response, '全部商品分类')
        self.assertContains(response, f'category={self.root.slug}')

    def test_search_limits_brand_panel_to_matching_products(self):
        Product.objects.create(
            name='同分类其他产品',
            style_code='STYLE-C-3',
            brand='其他品牌',
            category=self.leaf,
            status=PublishStatus.PUBLISHED,
        )
        response = self.client.get(
            reverse('catalog:product_list'),
            {'category': self.leaf.slug, 'q': '柱状活性炭'},
        )
        self.assertContains(response, '测试品牌')
        self.assertNotContains(response, '>其他品牌</a>')

    def test_frontend_product_list_hides_prices_and_price_sorting(self):
        response = self.client.get(reverse('catalog:product_list'))

        self.assertNotContains(response, '价格 &uarr;')
        self.assertNotContains(response, '价格 &darr;')
        self.assertNotContains(response, '价格待确认')


class QtyFilterTests(TestCase):
    def test_qty_filter_strips_decimal_zeros(self):
        from catalog.templatetags.catalog_extras import qty

        self.assertEqual(qty(Decimal('2.00')), '2')
        self.assertEqual(qty(Decimal('3')), '3')
        self.assertEqual(qty(Decimal('2.50')), '2.5')


class ProductVariantDisplayTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='活性炭', slug='active-carbon')
        CategoryAttribute.objects.create(
            category=self.category,
            name='孔径',
            code='pore_size',
            sort_order=10,
        )
        self.product = Product.objects.create(
            name='柱状活性炭',
            style_code='CARBON-SERIES',
            category=self.category,
            status=PublishStatus.PUBLISHED,
        )
        self.large = SKU.objects.create(
            product=self.product,
            internal_sku_code='CARBON-10',
            attributes={'pore_size': '10mm'},
            status=PublishStatus.PUBLISHED,
        )
        self.base = SKU.objects.create(
            product=self.product,
            internal_sku_code='CARBON-2',
            attributes={'pore_size': '2mm'},
            status=PublishStatus.PUBLISHED,
        )

    def test_product_list_uses_first_natural_attribute_value_as_base_sku(self):
        response = self.client.get(reverse('catalog:product_list'))

        self.assertEqual(response.context['products'][0].base_sku, self.base)
        self.assertContains(response, f'/cart/add/{self.base.id}/')

    def test_detail_defaults_to_base_sku_and_filters_other_models(self):
        url = reverse('catalog:product_detail', kwargs={'pk': self.product.pk})
        response = self.client.get(url, {'variant_pore_size': '10mm'})

        self.assertEqual(response.context['selected_sku'], self.base)
        self.assertEqual(response.context['variant_skus'], [self.large])
        self.assertContains(response, '其他所有型号')
        self.assertContains(response, '孔径')
        self.assertContains(response, '10mm')
        self.assertNotContains(response, '订货编码：CARBON-2')

    def test_detail_product_name_contains_selected_sku_attributes(self):
        url = reverse('catalog:product_detail', kwargs={'pk': self.product.pk})

        base_response = self.client.get(url)
        large_response = self.client.get(url, {'sku': self.large.pk})

        self.assertContains(base_response, '<span>2mm</span>', html=True)
        self.assertContains(large_response, '<span>10mm</span>', html=True)
        self.assertNotContains(base_response, '<span>孔径2mm</span>', html=True)

    def test_detail_hides_style_code_and_cas(self):
        self.product.cas_no = '7440-44-0'
        self.product.save(update_fields=['cas_no'])

        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertNotContains(response, '款式编码')
        self.assertNotContains(response, '7440-44-0')

    def test_detail_hides_generic_attribute_summary(self):
        self.base.sku_attribute_text = '2mm 通用属性汇总'
        self.base.save(update_fields=['sku_attribute_text'])

        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertNotContains(response, '<span>属性</span>', html=True)
        self.assertNotContains(response, '通用属性汇总')
        self.assertContains(response, '<span>孔径：</span>', html=True)

    def test_detail_hides_stock_summary(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertNotContains(response, '<span>库存</span>', html=True)

    def test_detail_attributes_use_colons_and_regular_value_weight(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertContains(response, 'class="sku-category-attribute"')
        self.assertContains(response, '<span>孔径：</span>', html=True)
        self.assertNotContains(response, '<strong>2mm</strong>', html=True)

    def test_detail_hides_moq_and_price_summary_items(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertNotContains(response, '<span>起订量</span>', html=True)
        self.assertNotContains(response, '<span>价格</span>', html=True)

    def test_variant_list_hides_base_model_badge(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertNotContains(response, '基础型号')

    def test_variant_list_uses_quantity_steppers(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertContains(response, 'quantity-stepper--compact', count=2)
        self.assertContains(response, 'data-quantity-change="-1"', count=3)
        self.assertContains(response, 'data-quantity-change="1"', count=3)

    def test_detail_removes_product_information_section(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertNotContains(response, '<h2>产品信息</h2>', html=True)
        self.assertNotContains(response, '<h3>类目参数</h3>', html=True)
        self.assertNotContains(response, '<h3>安全与运输</h3>', html=True)
        self.assertNotContains(response, '<h3>规格参数</h3>', html=True)
        self.assertNotContains(response, '<h3>简介</h3>', html=True)
        self.assertNotContains(response, '<h3>包装与储存</h3>', html=True)

    def test_detail_places_quantity_stepper_before_cart_button(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))
        content = response.content.decode()

        self.assertContains(response, 'class="quantity-stepper"')
        self.assertContains(response, 'data-quantity-change="-1"')
        self.assertContains(response, 'data-quantity-change="1"')
        self.assertContains(response, 'js/cart-add.js?v=20260712-cart-auto-update')
        self.assertLess(content.index('class="quantity-stepper"'), content.index('class="buy-actions-submit"'))

    def test_detail_hides_other_models_without_selectable_attributes(self):
        plain_category = Category.objects.create(name='普通分类', slug='plain-category')
        plain_product = Product.objects.create(
            name='普通产品',
            style_code='PLAIN-SERIES',
            category=plain_category,
            status=PublishStatus.PUBLISHED,
        )
        SKU.objects.create(
            product=plain_product,
            internal_sku_code='PLAIN-1',
            status=PublishStatus.PUBLISHED,
        )

        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': plain_product.pk}))

        self.assertNotContains(response, '其他所有型号')

    def test_detail_does_not_show_sku_select_dropdown(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertNotContains(response, '选择规格')
        self.assertNotContains(response, "window.location='?sku='")

    def test_variant_filters_refresh_immediately_without_apply_button(self):
        response = self.client.get(reverse('catalog:product_detail', kwargs={'pk': self.product.pk}))

        self.assertContains(response, 'data-variant-filter-form')
        self.assertContains(response, 'js/product-variants.js')
        self.assertNotContains(response, 'onchange="this.form.submit()"')
        self.assertNotContains(response, '应用筛选')

    def test_variant_list_paginates_at_five_items(self):
        for index in range(3, 8):
            SKU.objects.create(
                product=self.product,
                internal_sku_code=f'CARBON-{index}',
                attributes={'pore_size': f'{index}mm'},
                status=PublishStatus.PUBLISHED,
            )
        url = reverse('catalog:product_detail', kwargs={'pk': self.product.pk})

        first_page = self.client.get(url)
        second_page = self.client.get(url, {'variant_page': 2})

        self.assertEqual(len(first_page.context['variant_skus']), 5)
        self.assertEqual(len(second_page.context['variant_skus']), 2)
        self.assertContains(first_page, 'aria-label="系列型号分页"')
        self.assertContains(first_page, 'variant_page=2')


class ProductImageGalleryTests(TestCase):
    def setUp(self):
        self.product = Product.objects.create(
            name='测试工业清洁剂',
            style_code='STYLE-IMAGE-1',
            status=PublishStatus.PUBLISHED,
        )
        self.sku = SKU.objects.create(
            product=self.product,
            internal_sku_code='SKU-IMAGE-1',
            status=PublishStatus.PUBLISHED,
        )

    def make_image_tree(self, root):
        main_dir = root / self.product.style_code / '主图透图'
        detail_dir = root / self.product.style_code / '详情页'
        main_dir.mkdir(parents=True)
        detail_dir.mkdir()
        (main_dir / '主图.jpg').write_bytes(b'main-image')
        (main_dir / 'other.jpg').write_bytes(b'other-image')
        (detail_dir / '02.jpg').write_bytes(b'detail-two')
        (detail_dir / '01.jpg').write_bytes(b'detail-one')
        return main_dir, detail_dir

    def test_detail_uses_named_main_image_and_all_detail_images(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            self.make_image_tree(root)
            with override_settings(PRODUCT_IMAGE_ROOT=root):
                response = self.client.get(
                    reverse('catalog:product_detail', kwargs={'pk': self.product.pk})
                )

        self.assertContains(response, 'STYLE-IMAGE-1/%E4%B8%BB%E5%9B%BE%E9%80%8F%E5%9B%BE/%E4%B8%BB%E5%9B%BE.jpg')
        self.assertContains(response, '产品详情')
        self.assertContains(response, 'data-detail-image-open', count=2)
        self.assertContains(response, 'js/product-gallery.js')

    def test_product_list_uses_output_main_image_for_card_and_cart_animation(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            self.make_image_tree(root)
            with override_settings(PRODUCT_IMAGE_ROOT=root):
                response = self.client.get(reverse('catalog:product_list'))

        expected_url = (
            '/products/media/STYLE-IMAGE-1/'
            '%E4%B8%BB%E5%9B%BE%E9%80%8F%E5%9B%BE/'
            '%E4%B8%BB%E5%9B%BE.jpg'
        )
        self.assertContains(response, f'<img src="{expected_url}"', html=False)
        self.assertContains(response, f'data-image-url="{expected_url}"')
        self.assertNotContains(response, '/static/img/product-placeholder.jpg')

    def test_product_media_serves_only_gallery_images(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            main_dir, _ = self.make_image_tree(root)
            (main_dir / 'notes.txt').write_text('private note')
            with override_settings(PRODUCT_IMAGE_ROOT=root):
                image_response = self.client.get(
                    reverse(
                        'catalog:product_media',
                        kwargs={'asset_path': 'STYLE-IMAGE-1/主图透图/主图.jpg'},
                    )
                )
                text_response = self.client.get(
                    reverse(
                        'catalog:product_media',
                        kwargs={'asset_path': 'STYLE-IMAGE-1/主图透图/notes.txt'},
                    )
                )

                self.assertEqual(image_response.status_code, 200)
                self.assertEqual(b''.join(image_response.streaming_content), b'main-image')
                self.assertEqual(image_response['Cache-Control'], 'public, max-age=86400')
                self.assertEqual(text_response.status_code, 404)

    def test_selected_sku_uses_its_variant_image_directory(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            variant_main = (
                root
                / self.product.style_code
                / self.sku.internal_sku_code
                / '主图透图'
            )
            variant_main.mkdir(parents=True)
            (variant_main / '主图.png').write_bytes(b'variant-main')

            with override_settings(PRODUCT_IMAGE_ROOT=root):
                response = self.client.get(
                    reverse('catalog:product_detail', kwargs={'pk': self.product.pk})
                )

        self.assertContains(
            response,
            'STYLE-IMAGE-1/SKU-IMAGE-1/%E4%B8%BB%E5%9B%BE%E9%80%8F%E5%9B%BE/%E4%B8%BB%E5%9B%BE.png',
        )


class ProductImageAdminTests(TestCase):
    def setUp(self):
        self.product = Product.objects.create(
            name='后台图片测试产品',
            style_code='ADMIN-IMAGE-1',
            status=PublishStatus.PUBLISHED,
        )
        self.user = get_user_model().objects.create_superuser(
            'image-admin',
            password='pw38421844',
        )
        self.client.force_login(self.user)

    def image_admin_url(self):
        return reverse('admin:catalog_product_images', args=[self.product.pk])

    def test_product_change_page_links_to_image_manager(self):
        with TemporaryDirectory() as tmp, override_settings(PRODUCT_IMAGE_ROOT=Path(tmp)):
            response = self.client.get(
                reverse('admin:catalog_product_change', args=[self.product.pk])
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '上传和管理产品图片')
        self.assertContains(response, self.image_admin_url())
        self.assertNotContains(response, 'name="image"')
        self.assertNotContains(response, 'name="remote_image_url"')
        self.assertNotContains(response, 'name="image_source_url"')

    def test_admin_uploads_main_and_detail_images(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            with override_settings(PRODUCT_IMAGE_ROOT=root):
                main_response = self.client.post(
                    self.image_admin_url(),
                    {
                        'action': 'upload',
                        'image_type': 'main',
                        'images': [
                            SimpleUploadedFile('front.png', b'front', content_type='image/png'),
                            SimpleUploadedFile('side.jpg', b'side', content_type='image/jpeg'),
                        ],
                    },
                )
                detail_response = self.client.post(
                    self.image_admin_url(),
                    {
                        'action': 'upload',
                        'image_type': 'detail',
                        'images': SimpleUploadedFile(
                            'details.webp',
                            b'details',
                            content_type='image/webp',
                        ),
                    },
                )
                page = self.client.get(self.image_admin_url())

            product_root = root / self.product.style_code
            self.assertRedirects(main_response, self.image_admin_url())
            self.assertRedirects(detail_response, self.image_admin_url())
            self.assertEqual((product_root / '主图透图' / '主图.png').read_bytes(), b'front')
            self.assertEqual((product_root / '主图透图' / 'side.jpg').read_bytes(), b'side')
            self.assertEqual((product_root / '详情页' / 'details.webp').read_bytes(), b'details')
            self.assertContains(page, '当前产品主图')
            self.assertContains(page, 'details.webp')

    def test_admin_can_promote_and_delete_images(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            main_dir = root / self.product.style_code / '主图透图'
            detail_dir = root / self.product.style_code / '详情页'
            main_dir.mkdir(parents=True)
            detail_dir.mkdir()
            (main_dir / '主图.png').write_bytes(b'old-main')
            (main_dir / 'new.jpg').write_bytes(b'new-main')
            (detail_dir / 'delete-me.png').write_bytes(b'detail')

            with override_settings(PRODUCT_IMAGE_ROOT=root):
                promote_response = self.client.post(
                    self.image_admin_url(),
                    {
                        'action': 'set_main',
                        'image_path': '主图透图/new.jpg',
                    },
                )
                delete_response = self.client.post(
                    self.image_admin_url(),
                    {
                        'action': 'delete',
                        'image_path': '详情页/delete-me.png',
                    },
                )

            self.assertRedirects(promote_response, self.image_admin_url())
            self.assertRedirects(delete_response, self.image_admin_url())
            self.assertEqual((main_dir / '主图.jpg').read_bytes(), b'new-main')
            self.assertEqual((main_dir / '原主图.png').read_bytes(), b'old-main')
            self.assertFalse((detail_dir / 'delete-me.png').exists())

    def test_admin_rejects_paths_outside_current_product(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp)
            other_image = root / 'OTHER' / '详情页' / 'secret.jpg'
            other_image.parent.mkdir(parents=True)
            other_image.write_bytes(b'keep')

            with override_settings(PRODUCT_IMAGE_ROOT=root):
                response = self.client.post(
                    self.image_admin_url(),
                    {
                        'action': 'delete',
                        'image_path': '../OTHER/详情页/secret.jpg',
                    },
                )

            self.assertEqual(response.status_code, 404)
            self.assertTrue(other_image.exists())

    def test_admin_rejects_non_image_upload(self):
        with TemporaryDirectory() as tmp, override_settings(PRODUCT_IMAGE_ROOT=Path(tmp)):
            response = self.client.post(
                self.image_admin_url(),
                {
                    'action': 'upload',
                    'image_type': 'main',
                    'images': SimpleUploadedFile(
                        'notes.txt',
                        b'not an image',
                        content_type='text/plain',
                    ),
                },
                follow=True,
            )

        self.assertContains(response, '仅支持')


class ProductImageArchiveTests(TestCase):
    def setUp(self):
        self.first = Product.objects.create(
            name='批量图片产品一',
            style_code='BULK-IMAGE-1',
            status=PublishStatus.PUBLISHED,
        )
        self.second = Product.objects.create(
            name='批量图片产品二',
            style_code='BULK-IMAGE-2',
            status=PublishStatus.PUBLISHED,
        )

    def make_archive(self, files):
        buffer = BytesIO()
        with ZipFile(buffer, 'w') as archive:
            for path, content in files.items():
                archive.writestr(path, content)
        buffer.seek(0)
        return buffer

    def test_archive_completely_replaces_duplicate_style_directories(self):
        from catalog.product_image_archive import import_product_image_archive

        with TemporaryDirectory() as tmp:
            root = Path(tmp) / 'product_catalog'
            old_main = root / self.first.style_code / '主图透图'
            old_main.mkdir(parents=True)
            (old_main / 'old.jpg').write_bytes(b'old')
            (root / self.first.style_code / 'stale.txt').write_text('remove me')
            archive = self.make_archive(
                {
                    f'{self.first.style_code}/主图透图/主图.png': b'new-main',
                    f'{self.first.style_code}/详情页/01.webp': b'new-detail',
                    f'{self.second.style_code}/主图透图/主图.jpg': b'second-main',
                }
            )

            result = import_product_image_archive(
                archive,
                known_style_codes=[self.first.style_code, self.second.style_code],
                image_root=root,
            )

            self.assertEqual(result.image_count, 3)
            self.assertEqual(result.style_codes, ('BULK-IMAGE-1', 'BULK-IMAGE-2'))
            self.assertFalse((root / self.first.style_code / '主图透图' / 'old.jpg').exists())
            self.assertFalse((root / self.first.style_code / 'stale.txt').exists())
            self.assertEqual(
                (root / self.first.style_code / '主图透图' / '主图.png').read_bytes(),
                b'new-main',
            )
            self.assertEqual(
                (root / self.second.style_code / '主图透图' / '主图.jpg').read_bytes(),
                b'second-main',
            )

    def test_archive_accepts_product_catalog_wrapper_directory(self):
        from catalog.product_image_archive import import_product_image_archive

        with TemporaryDirectory() as tmp:
            root = Path(tmp) / 'product_catalog'
            archive = self.make_archive(
                {
                    f'product_catalog/{self.first.style_code}/主图透图/主图.avif': b'avif',
                }
            )

            result = import_product_image_archive(
                archive,
                known_style_codes=[self.first.style_code],
                image_root=root,
            )

            self.assertEqual(result.style_codes, (self.first.style_code,))
            self.assertTrue(
                (root / self.first.style_code / '主图透图' / '主图.avif').is_file()
            )

    def test_archive_rejects_unknown_style_without_changing_existing_images(self):
        from catalog.product_image_archive import (
            ProductImageArchiveError,
            import_product_image_archive,
        )

        with TemporaryDirectory() as tmp:
            root = Path(tmp) / 'product_catalog'
            existing = root / self.first.style_code / '主图透图' / '主图.jpg'
            existing.parent.mkdir(parents=True)
            existing.write_bytes(b'keep')
            archive = self.make_archive(
                {'UNKNOWN-STYLE/主图透图/主图.jpg': b'rejected'}
            )

            with self.assertRaisesRegex(ProductImageArchiveError, '款式编码不存在'):
                import_product_image_archive(
                    archive,
                    known_style_codes=[self.first.style_code],
                    image_root=root,
                )

            self.assertEqual(existing.read_bytes(), b'keep')

    def test_archive_rejects_path_traversal(self):
        from catalog.product_image_archive import (
            ProductImageArchiveError,
            import_product_image_archive,
        )

        archive = self.make_archive(
            {f'{self.first.style_code}/主图透图/../../../outside.jpg': b'bad'}
        )
        with TemporaryDirectory() as tmp:
            root = Path(tmp) / 'product_catalog'
            with self.assertRaisesRegex(ProductImageArchiveError, '不安全路径'):
                import_product_image_archive(
                    archive,
                    known_style_codes=[self.first.style_code],
                    image_root=root,
                )
            self.assertFalse((Path(tmp) / 'outside.jpg').exists())

    def test_directory_replacement_rolls_back_previous_styles_on_failure(self):
        from catalog.product_image_archive import (
            ProductImageArchiveError,
            replace_style_directories,
        )

        with TemporaryDirectory() as tmp:
            base = Path(tmp)
            root = base / 'product_catalog'
            staging = base / 'staging'
            work = base / 'work'
            root.mkdir()
            staging.mkdir()
            work.mkdir()
            old = root / self.first.style_code / '主图透图'
            old.mkdir(parents=True)
            (old / '主图.jpg').write_bytes(b'old')
            new = staging / self.first.style_code / '主图透图'
            new.mkdir(parents=True)
            (new / '主图.jpg').write_bytes(b'new')

            with self.assertRaisesRegex(ProductImageArchiveError, '无法准备款式目录'):
                replace_style_directories(
                    staging,
                    root,
                    {self.first.style_code, self.second.style_code},
                    work,
                )

            self.assertEqual(
                (root / self.first.style_code / '主图透图' / '主图.jpg').read_bytes(),
                b'old',
            )


class ProductImageArchiveAdminTests(TestCase):
    def setUp(self):
        self.product = Product.objects.create(
            name='后台批量图片产品',
            style_code='ADMIN-BULK-1',
            status=PublishStatus.PUBLISHED,
        )
        self.user = get_user_model().objects.create_superuser(
            'bulk-image-admin',
            password='pw38421844',
        )
        self.client.force_login(self.user)

    def make_uploaded_archive(self):
        buffer = BytesIO()
        with ZipFile(buffer, 'w') as archive:
            archive.writestr(
                f'{self.product.style_code}/主图透图/主图.png',
                b'bulk-main',
            )
        return SimpleUploadedFile(
            'product-images.zip',
            buffer.getvalue(),
            content_type='application/zip',
        )

    def test_product_list_links_to_bulk_image_upload(self):
        response = self.client.get(reverse('admin:catalog_product_changelist'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '批量上传产品图片')
        self.assertContains(response, reverse('admin:catalog_product_bulk_images'))

    def test_admin_bulk_upload_requires_confirmation_then_imports_archive(self):
        with TemporaryDirectory() as tmp:
            root = Path(tmp) / 'product_catalog'
            with override_settings(PRODUCT_IMAGE_ROOT=root):
                rejected = self.client.post(
                    reverse('admin:catalog_product_bulk_images'),
                    {'archive': self.make_uploaded_archive()},
                    follow=True,
                )
                imported = self.client.post(
                    reverse('admin:catalog_product_bulk_images'),
                    {
                        'archive': self.make_uploaded_archive(),
                        'confirm_replace': 'on',
                    },
                    follow=True,
                )

            self.assertContains(rejected, '请确认同款式编码目录将被完全覆盖')
            self.assertContains(imported, '批量上传完成')
            self.assertEqual(
                (root / self.product.style_code / '主图透图' / '主图.png').read_bytes(),
                b'bulk-main',
            )


class CategoryImportExportTests(TestCase):
    def setUp(self):
        self.root = Category.objects.create(name='车间化学品', slug='workshop', sort_order=10)
        self.mid = Category.objects.create(name='活性炭', slug='carbon', parent=self.root, sort_order=20)
        CategoryAttribute.objects.create(
            category=self.mid,
            name='碘值',
            code='iodine_value',
            data_type='number',
            sort_order=10,
        )

    def test_export_and_import_roundtrip(self):
        from io import BytesIO

        from catalog.category_io import export_categories_workbook, import_categories_workbook

        workbook = export_categories_workbook(Category.objects.all())
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        from openpyxl import load_workbook

        imported = load_workbook(buffer, data_only=True)
        result = import_categories_workbook(imported)
        self.assertEqual(result.failures, [])
        self.assertEqual(Category.objects.filter(slug='workshop').count(), 1)
        self.assertEqual(Category.objects.filter(slug='carbon').count(), 1)
        self.assertTrue(CategoryAttribute.objects.filter(category__slug='carbon', code='iodine_value').exists())

    def test_admin_export_all_categories(self):
        user = get_user_model().objects.create_superuser('admin', password='pw38421844')
        client = Client()
        client.force_login(user)
        response = client.get(reverse('admin:catalog_category_export_all'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_hierarchical_import_without_slug(self):
        from io import BytesIO

        from catalog.category_io import import_categories_workbook
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '总目录'
        sheet.append(['一级目录', '二级目录', '三级目录', '说明'])
        sheet.append(['车间化学品', '活性炭', '柱状活性炭', ''])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        from openpyxl import load_workbook

        result = import_categories_workbook(load_workbook(buffer, data_only=True))
        self.assertEqual(result.failures, [])
        self.assertEqual(Category.objects.filter(name='柱状活性炭').count(), 1)
        leaf = Category.objects.get(name='柱状活性炭')
        self.assertEqual(leaf.parent.name, '活性炭')
        self.assertTrue(leaf.slug)


class CategoryAdminDisplayTests(TestCase):
    def test_category_tree_choices_and_labels(self):
        from catalog.admin import category_choice_label, category_tree_choices

        root = Category.objects.create(name='润滑', slug='lube', sort_order=10)
        mid = Category.objects.create(name='车用润滑油', slug='lube-auto', parent=root, sort_order=20)
        leaf = Category.objects.create(name='车用齿轮油', slug='lube-auto-gear', parent=mid, sort_order=30)

        ordered = category_tree_choices(Category.objects.all())
        self.assertEqual([item.pk for item in ordered], [root.pk, mid.pk, leaf.pk])
        self.assertEqual(category_choice_label(root), '润滑')
        self.assertEqual(category_choice_label(mid), '\u3000车用润滑油')
        self.assertEqual(category_choice_label(leaf), '\u3000\u3000车用齿轮油')


class ProductImportExportTests(TestCase):
    def setUp(self):
        self.category = Category.objects.create(name='测试分类', slug='test-category')
        self.attribute = CategoryAttribute.objects.create(
            category=self.category,
            name='孔径',
            code='pore_size',
            is_required=True,
        )
        self.product = Product.objects.create(
            name='测试树脂',
            style_code='STYLE-1',
            brand='测试品牌',
            category=self.category,
            status=PublishStatus.PUBLISHED,
        )
        self.sku = SKU.objects.create(
            product=self.product,
            internal_sku_code='SKU-1',
            jst_sku_id='SKU-1',
            price=Decimal('12.50'),
            moq=Decimal('2'),
            order_step=Decimal('1'),
            attributes={'pore_size': '4mm'},
            status=PublishStatus.PUBLISHED,
        )

    def test_export_and_import_roundtrip(self):
        from io import BytesIO

        from catalog.product_io import export_products_workbook, import_products_workbook
        from openpyxl import load_workbook

        workbook = export_products_workbook(Product.objects.all())
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        result = import_products_workbook(load_workbook(buffer, data_only=True), allow_missing_price=True)
        self.assertEqual(result.failures, [])
        self.assertEqual(Product.objects.filter(style_code='STYLE-1').count(), 1)
        self.assertEqual(SKU.objects.filter(internal_sku_code='SKU-1').count(), 1)
        self.assertEqual(SKU.objects.get(internal_sku_code='SKU-1').attributes, {'pore_size': '4mm'})

    def test_import_without_special_attribute_column_preserves_existing_value(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '产品'
        sheet.append(['SKU', '产品名称', '产品型号', '销售价格', '分类'])
        sheet.append(['SKU-1', '测试树脂', 'STYLE-1', '12.50', '测试分类'])

        from catalog.product_io import import_products_workbook

        result = import_products_workbook(workbook)

        self.assertEqual(result.failures, [])
        self.assertEqual(SKU.objects.get(pk=self.sku.pk).attributes, {'pore_size': '4mm'})

    def test_admin_form_modifies_special_attributes(self):
        from catalog.admin import SKUAdminForm

        data = {
            field.name: getattr(self.sku, field.name)
            for field in SKU._meta.fields
            if field.name not in {'id', 'created_at', 'updated_at'}
        }
        data['product'] = self.product.pk
        data.pop('attributes', None)
        data['attributes__pore_size'] = '__new__'
        data['attributes__pore_size__new'] = '8mm'
        data['source_raw_row'] = '{}'
        form = SKUAdminForm(data=data, instance=self.sku)

        self.assertTrue(form.is_valid(), form.errors)
        form.save()
        self.assertEqual(SKU.objects.get(pk=self.sku.pk).attributes, {'pore_size': '8mm'})

    def test_admin_form_renders_named_special_attribute_inputs(self):
        from catalog.admin import SKUAdminForm

        form = SKUAdminForm(instance=self.sku)
        html = str(form['attributes'])

        self.assertIn('孔径', html)
        self.assertIn('<select', html)
        self.assertIn('name="attributes__pore_size"', html)
        self.assertIn('<option value="4mm" selected>4mm</option>', html)
        self.assertIn('＋ 新增值…', html)
        self.assertIn('name="attributes__pore_size__new"', html)
        self.assertNotIn('textarea', html)

    def test_product_admin_inline_hides_legacy_quantity_and_spec_fields(self):
        from catalog.admin import SKUInline

        self.assertNotIn('sku_attribute_text', SKUInline.fields)
        self.assertNotIn('moq', SKUInline.fields)
        self.assertNotIn('order_step', SKUInline.fields)
        self.assertIn('attributes', SKUInline.fields)

    def test_large_product_admin_uses_separate_sku_management(self):
        user = get_user_model().objects.create_superuser('large-admin', password='pw38421844')
        extra_skus = [
            SKU(
                product=self.product,
                internal_sku_code=f'LARGE-SKU-{index}',
                attributes={'pore_size': f'{index}mm'},
            )
            for index in range(211)
        ]
        SKU.objects.bulk_create(extra_skus)
        client = Client()
        client.force_login(user)

        response = client.get(reverse('admin:catalog_product_change', args=[self.product.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="skus-TOTAL_FORMS"')
        self.assertContains(response, '管理该系列的 212 个具体型号')
        self.assertContains(response, f'product__id__exact={self.product.pk}')
        self.assertLess(response.content.count(b'<input'), 1000)

        sku_list = client.get(
            reverse('admin:catalog_sku_changelist'),
            {'product__id__exact': self.product.pk},
        )
        self.assertEqual(sku_list.status_code, 200)
        self.assertEqual(sku_list.context['cl'].result_count, 212)

    def test_single_sku_is_editable_on_product_admin_page(self):
        user = get_user_model().objects.create_superuser('single-sku-admin', password='pw38421844')
        client = Client()
        client.force_login(user)

        response = client.get(reverse('admin:catalog_product_change', args=[self.product.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="skus-TOTAL_FORMS"')
        self.assertContains(response, self.sku.internal_sku_code)
        self.assertNotContains(response, '打开完整 SKU 编辑页')
        self.assertContains(response, 'name="single_sku_attributes__pore_size"')
        self.assertContains(response, 'name="single_sku_code"')
        self.assertContains(response, '产品图片资源')
        self.assertContains(response, '上传和管理产品图片')
        self.assertContains(
            response,
            reverse('admin:catalog_product_images', args=[self.product.pk]),
        )

    def test_product_admin_saves_single_sku_fields(self):
        from catalog.admin import ProductAdminForm

        data = {
            field.name: getattr(self.product, field.name)
            for field in Product._meta.fields
            if field.name not in {'id', 'created_at', 'updated_at', 'image'}
        }
        data['category'] = self.category.pk
        data['single_sku_code'] = 'SKU-1-UPDATED'
        data['single_sku_price'] = '18.50'
        data['single_sku_stock_status'] = self.sku.stock_status
        data['single_sku_attributes__pore_size'] = '__new__'
        data['single_sku_attributes__pore_size__new'] = '10mm'
        data['single_sku_status'] = self.sku.status
        form = ProductAdminForm(data=data, instance=self.product)

        self.assertTrue(form.is_valid(), form.errors)
        form.save()
        form.save_single_sku()
        self.sku.refresh_from_db()
        self.assertEqual(self.sku.internal_sku_code, 'SKU-1-UPDATED')
        self.assertEqual(self.sku.price, Decimal('18.50'))
        self.assertEqual(self.sku.attributes, {'pore_size': '10mm'})

    def test_sku_admin_links_back_to_product_series(self):
        user = get_user_model().objects.create_superuser('sku-admin', password='pw38421844')
        client = Client()
        client.force_login(user)

        response = client.get(reverse('admin:catalog_sku_change', args=[self.sku.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '返回产品系列：')
        self.assertContains(response, reverse('admin:catalog_product_change', args=[self.product.pk]))

    def test_admin_export_all_products(self):
        user = get_user_model().objects.create_superuser('admin', password='pw38421844')
        client = Client()
        client.force_login(user)
        response = client.get(reverse('admin:catalog_product_export_all'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
