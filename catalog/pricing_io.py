"""客户协议价批量导入/导出。

字段格式：
  公司名称 | 内部 SKU 编码 | 协议价 | 起订量 | 封顶量 | 生效开始 | 生效结束 | 启用 | 备注
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from customers.models import Customer

from .models import CustomerSKUPrice, PriceHistory, SKU


PRICE_SHEET = '客户协议价'

EXPORT_HEADERS = (
    '公司名称',
    '内部 SKU 编码',
    '协议价',
    '起订量',
    '封顶量',
    '生效开始',
    '生效结束',
    '启用',
    '备注',
    '最后更新',
)


@dataclass
class CustomerPriceImportResult:
    created: int = 0
    updated: int = 0
    disabled: int = 0
    failures: list[str] = field(default_factory=list)


def _parse_decimal(text, default=None):
    if text in (None, '', ' '):
        return default
    try:
        return Decimal(str(text).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f'数值格式不正确：{text!r}')


def _parse_date(text):
    if text in (None, ''):
        return None
    if isinstance(text, date):
        return text
    text = str(text).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, '%Y-%m-%d').date()
    except ValueError:
        try:
            return datetime.strptime(text, '%Y/%m/%d').date()
        except ValueError as error:
            raise ValueError(f'日期格式应为 YYYY-MM-DD：{text!r}') from error


def _parse_bool(text):
    if text in (None, ''):
        return True
    text = str(text).strip()
    if text in {'是', '启用', 'Y', 'y', 'yes', 'YES', '1', 'true', 'TRUE'}:
        return True
    if text in {'否', '停用', 'N', 'n', 'no', 'NO', '0', 'false', 'FALSE'}:
        return False
    raise ValueError(f'启用/停用字段无法识别：{text!r}')


def _row_text(row, idx):
    if idx >= len(row):
        return ''
    value = row[idx]
    return '' if value is None else str(value).strip()


def export_customer_prices_workbook(queryset):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = PRICE_SHEET
    worksheet.append(list(EXPORT_HEADERS))

    records = queryset.select_related('customer', 'sku').order_by('customer__company_name', 'sku__internal_sku_code')
    for record in records:
        worksheet.append([
            record.customer.company_name,
            record.sku.internal_sku_code,
            record.price,
            record.min_qty or '',
            record.max_qty or '',
            record.valid_from.isoformat() if record.valid_from else '',
            record.valid_to.isoformat() if record.valid_to else '',
            '启用' if record.is_active else '停用',
            record.remark,
            timezone.localtime(record.updated_at).strftime('%Y-%m-%d %H:%M'),
        ])
    return workbook


def export_customer_prices_response(queryset, filename_prefix='customer-prices'):
    workbook = export_customer_prices_workbook(queryset)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    timestamp = timezone.localtime().strftime('%Y%m%d-%H%M%S')
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}-{timestamp}.xlsx"'
    return response


def _resolve_customer(text):
    name = (text or '').strip()
    if not name:
        raise ValueError('公司名称为空')
    customer = Customer.objects.filter(company_name=name).first()
    if not customer:
        raise ValueError(f'找不到客户：{name}')
    return customer


def _resolve_sku(text):
    code = (text or '').strip()
    if not code:
        raise ValueError('内部 SKU 编码为空')
    sku = SKU.objects.filter(internal_sku_code=code).first()
    if not sku:
        raise ValueError(f'找不到 SKU：{code}')
    return sku


@transaction.atomic
def import_customer_prices_workbook(workbook, *, operator=None, dry_run=False):
    result = CustomerPriceImportResult()
    if PRICE_SHEET not in workbook.sheetnames:
        raise ValueError(f'Excel 中缺少工作表「{PRICE_SHEET}」')
    worksheet = workbook[PRICE_SHEET]
    headers = [str(cell.value or '').strip() for cell in worksheet[1]]
    expected = list(EXPORT_HEADERS[:8])
    for idx, name in enumerate(expected):
        if idx >= len(headers) or headers[idx] != name:
            raise ValueError(f'第 {idx + 1} 列表头应为「{name}」，当前为「{headers[idx] if idx < len(headers) else ""}」')

    sid = transaction.savepoint()
    try:
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell not in (None, '') for cell in row):
                continue
            try:
                customer = _resolve_customer(_row_text(row, 0))
                sku = _resolve_sku(_row_text(row, 1))
                price = _parse_decimal(row[2])
                if price is None:
                    raise ValueError('协议价不能为空')
                min_qty = _parse_decimal(row[3])
                max_qty = _parse_decimal(row[4])
                valid_from = _parse_date(row[5])
                valid_to = _parse_date(row[6])
                is_active = _parse_bool(row[7])
                remark = _row_text(row, 8) if len(row) > 8 else ''

                if valid_from and valid_to and valid_to < valid_from:
                    raise ValueError('生效结束早于生效开始')

                existing = (
                    CustomerSKUPrice.objects
                    .filter(customer=customer, sku=sku, min_qty=min_qty, max_qty=max_qty)
                    .first()
                )
                if existing:
                    old_price = existing.price
                    existing.price = price
                    existing.valid_from = valid_from
                    existing.valid_to = valid_to
                    existing.is_active = is_active
                    existing.remark = remark
                    existing.save()
                    PriceHistory.objects.create(
                        customer=customer,
                        sku=sku,
                        price_record=existing,
                        change_type=PriceHistory.ChangeType.UPDATE,
                        old_price=old_price,
                        new_price=price,
                        operator=operator,
                    )
                    result.updated += 1
                else:
                    record = CustomerSKUPrice.objects.create(
                        customer=customer,
                        sku=sku,
                        price=price,
                        min_qty=min_qty,
                        max_qty=max_qty,
                        valid_from=valid_from,
                        valid_to=valid_to,
                        is_active=is_active,
                        remark=remark,
                    )
                    PriceHistory.objects.create(
                        customer=customer,
                        sku=sku,
                        price_record=record,
                        change_type=PriceHistory.ChangeType.CREATE,
                        old_price=None,
                        new_price=price,
                        operator=operator,
                    )
                    result.created += 1
            except Exception as exc:
                result.failures.append(f'第 {row_index} 行：{exc}')
        if dry_run:
            transaction.savepoint_rollback(sid)
        else:
            transaction.savepoint_commit(sid)
    except Exception:
        transaction.savepoint_rollback(sid)
        raise
    return result


def import_customer_prices_file(uploaded_file, *, operator=None, dry_run=False):
    workbook = load_workbook(uploaded_file, data_only=True)
    return import_customer_prices_workbook(workbook, operator=operator, dry_run=dry_run)
