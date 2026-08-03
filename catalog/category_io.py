from __future__ import annotations

from dataclasses import dataclass, field
from hashlib import md5
from io import BytesIO

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook

from .models import AttributeDataType, Category, CategoryAttribute


CATEGORY_SHEET = '分类'
CATEGORY_SHEET_ALIASES = (CATEGORY_SHEET, '总目录')
ATTRIBUTE_SHEET = '分类属性'

HIERARCHICAL_CATEGORY_HEADERS = ('一级目录', '二级目录', '三级目录', '说明')
SLUG_CATEGORY_HEADERS = ('slug', '分类名称', '上级slug', '说明')

HIERARCHICAL_ATTRIBUTE_HEADERS = (
    '一级目录',
    '二级目录',
    '三级目录',
    '属性名称',
    '属性编码',
    '数据类型',
    '必填',
    '用于筛选',
    '列表显示',
    '详情显示',
)
SLUG_ATTRIBUTE_HEADERS = (
    '分类slug',
    '属性名称',
    '属性编码',
    '数据类型',
    '必填',
    '用于筛选',
    '列表显示',
    '详情显示',
)

ATTRIBUTE_CODE_MAP = {
    'SKU': 'sku',
    '品牌': 'brand',
    '产品名称': 'product_name',
    '产品型号': 'model',
    '产品规格': 'spec',
    '销售价格': 'price',
    '颜色': 'color',
    '包材': 'package_material',
    '克重': 'gram_weight',
    '箱规': 'carton_spec',
    '碘值': 'iodine_value',
    '尺寸': 'size',
    '孔径': 'pore_size',
    '亚兰值': 'methylene_blue_value',
}
NUMERIC_ATTRIBUTES = {'克重', '碘值', '亚兰值'}

DATA_TYPE_LABELS = {choice.value: choice.label for choice in AttributeDataType}
DATA_TYPE_BY_LABEL = {choice.label: choice.value for choice in AttributeDataType}


@dataclass
class CategoryImportResult:
    created_categories: int = 0
    updated_categories: int = 0
    created_attributes: int = 0
    updated_attributes: int = 0
    failures: list[str] = field(default_factory=list)


def clean_text(value):
    if value is None:
        return ''
    return str(value).strip()


def stable_slug(*parts):
    label = '-'.join(part for part in parts if part)
    ascii_slug = slugify(label)
    if ascii_slug:
        return ascii_slug[:140]
    digest = md5(label.encode('utf-8')).hexdigest()[:12]
    return f'cat-{digest}'


def attribute_code(name):
    mapped = ATTRIBUTE_CODE_MAP.get(name)
    if mapped:
        return mapped
    ascii_code = slugify(name).replace('-', '_')
    if ascii_code:
        return ascii_code[:100]
    digest = md5(name.encode('utf-8')).hexdigest()[:12]
    return f'attr_{digest}'


def attribute_type(name):
    if name == '销售价格':
        return AttributeDataType.PRICE
    if name in NUMERIC_ATTRIBUTES:
        return AttributeDataType.NUMBER
    return AttributeDataType.TEXT


def category_path_key(parts):
    return '/'.join(part for part in parts if part)


def parse_bool(value, default=True):
    text = clean_text(value).lower()
    if not text:
        return default
    if text in {'1', 'true', 'yes', 'y', '是', '启用', 'enabled'}:
        return True
    if text in {'0', 'false', 'no', 'n', '否', '停用', 'disabled'}:
        return False
    raise ValueError(f'无法识别的布尔值：{value}')


def parse_int(value, default=0):
    text = clean_text(value)
    if not text:
        return default
    return int(float(text))


def category_sheet_name(workbook):
    for name in CATEGORY_SHEET_ALIASES:
        if name in workbook.sheetnames:
            return name
    return None


def breadcrumb_parts(category):
    return [item.name for item in category.breadcrumb()]


def expand_category_queryset(queryset):
    category_ids = set()
    for category in queryset.select_related('parent'):
        for item in category.breadcrumb():
            category_ids.add(item.pk)
        category_ids.update(child.pk for child in category.descendants())
    return Category.objects.filter(pk__in=category_ids).order_by('sort_order', 'name')


def export_categories_workbook(categories):
    workbook = Workbook()
    category_sheet = workbook.active
    category_sheet.title = CATEGORY_SHEET
    category_sheet.append(list(HIERARCHICAL_CATEGORY_HEADERS))

    categories = list(categories.select_related('parent'))
    for category in categories:
        parts = breadcrumb_parts(category)
        while len(parts) < 3:
            parts.append('')
        category_sheet.append(
            [
                parts[0],
                parts[1],
                parts[2],
                category.description,
            ]
        )

    attribute_sheet = workbook.create_sheet(ATTRIBUTE_SHEET)
    attribute_sheet.append(list(HIERARCHICAL_ATTRIBUTE_HEADERS))
    attributes = (
        CategoryAttribute.objects.filter(category__in=categories)
        .select_related('category', 'category__parent', 'category__parent__parent')
        .order_by('category__sort_order', 'category__name', 'sort_order', 'name')
    )
    for attribute in attributes:
        parts = breadcrumb_parts(attribute.category)
        while len(parts) < 3:
            parts.append('')
        attribute_sheet.append(
            [
                parts[0],
                parts[1],
                parts[2],
                attribute.name,
                attribute.code,
                DATA_TYPE_LABELS.get(attribute.data_type, attribute.data_type),
                '是' if attribute.is_required else '否',
                '是' if attribute.is_filterable else '否',
                '是' if attribute.is_list_visible else '否',
                '是' if attribute.is_detail_visible else '否',
            ]
        )

    return workbook


def export_categories_response(categories, filename_prefix='categories'):
    workbook = export_categories_workbook(categories)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    timestamp = timezone.localtime().strftime('%Y%m%d-%H%M%S')
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}-{timestamp}.xlsx"'
    return response


def _cell(row, header_index, header):
    index = header_index.get(header)
    if index is None or index >= len(row):
        return ''
    return clean_text(row[index])


def _parse_hierarchical_category_rows(worksheet, result):
    headers = [clean_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {header: index for index, header in enumerate(headers)}
    missing_headers = [header for header in ('一级目录', '二级目录', '三级目录') if header not in header_index]
    if missing_headers:
        result.failures.append(f'{worksheet.title} 缺少表头：{", ".join(missing_headers)}')
        return []

    category_rows = []
    seen_slugs = set()
    for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        level1 = _cell(row, header_index, '一级目录')
        level2 = _cell(row, header_index, '二级目录')
        level3 = _cell(row, header_index, '三级目录')
        if not any((level1, level2, level3)):
            continue
        if not level1:
            result.failures.append(f'{worksheet.title} 第 {row_no} 行：一级目录不能为空')
            continue
        description = _cell(row, header_index, '说明')

        parent_slug = ''
        path = []
        for depth, name in enumerate((level1, level2, level3), start=1):
            if not name:
                continue
            path.append(name)
            slug = stable_slug(*path)
            if slug in seen_slugs:
                parent_slug = slug
                continue
            seen_slugs.add(slug)
            category_rows.append(
                {
                    'row_no': row_no,
                    'slug': slug,
                    'name': name,
                    'parent_slug': parent_slug,
                    'sort_order': row_no * 10 + depth,
                    'is_active': True,
                    'description': description if depth == len([n for n in (level1, level2, level3) if n]) else '',
                    'path_key': category_path_key(path),
                }
            )
            parent_slug = slug
    return category_rows


def _parse_slug_category_rows(worksheet, result):
    headers = [clean_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {header: index for index, header in enumerate(headers)}
    missing_headers = [header for header in ('分类名称',) if header not in header_index]
    if missing_headers:
        result.failures.append(f'{worksheet.title} 缺少表头：{", ".join(missing_headers)}')
        return []

    category_rows = []
    for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        slug = _cell(row, header_index, 'slug')
        name = _cell(row, header_index, '分类名称')
        parent_slug = _cell(row, header_index, '上级slug')
        description = _cell(row, header_index, '说明')
        if not any((slug, name, parent_slug, description)):
            continue
        if not name:
            result.failures.append(f'{worksheet.title} 第 {row_no} 行：分类名称不能为空')
            continue
        if not slug:
            slug = stable_slug(name) if not parent_slug else stable_slug(parent_slug, name)
        category_rows.append(
            {
                'row_no': row_no,
                'slug': slug,
                'name': name,
                'parent_slug': parent_slug,
                'sort_order': row_no * 10,
                'is_active': True,
                'description': description,
                'path_key': '',
            }
        )
    return category_rows


def _parse_category_rows(workbook, result):
    sheet_name = category_sheet_name(workbook)
    if not sheet_name:
        result.failures.append(f'缺少工作表：{" 或 ".join(CATEGORY_SHEET_ALIASES)}')
        return []

    worksheet = workbook[sheet_name]
    headers = [clean_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if '一级目录' in headers:
        return _parse_hierarchical_category_rows(worksheet, result)
    return _parse_slug_category_rows(worksheet, result)


def _upsert_categories(category_rows, result):
    slug_map = {category.slug: category for category in Category.objects.select_related('parent')}
    path_map = {}
    for category in slug_map.values():
        path_map[category_path_key(breadcrumb_parts(category))] = category

    pending_rows = sorted(category_rows, key=lambda item: (item['parent_slug'], item['slug']))
    while pending_rows:
        next_pending = []
        progress = False
        for item in pending_rows:
            if item['parent_slug'] and item['parent_slug'] not in slug_map:
                next_pending.append(item)
                continue
            parent = slug_map.get(item['parent_slug']) if item['parent_slug'] else None
            existing = None
            if item.get('path_key'):
                existing = path_map.get(item['path_key'])
            if existing is None:
                existing = Category.objects.filter(parent=parent, name=item['name']).first()
            generated_slug = item['slug']
            slug = existing.slug if existing else generated_slug
            category, created = Category.objects.update_or_create(
                slug=slug,
                defaults={
                    'name': item['name'],
                    'parent': parent,
                    'sort_order': item['sort_order'],
                    'is_active': item['is_active'],
                    'description': item['description'],
                },
            )
            slug_map[category.slug] = category
            slug_map[generated_slug] = category
            item['slug'] = category.slug
            path_key = item['path_key'] or category_path_key(breadcrumb_parts(category))
            path_map[path_key] = category
            if created:
                result.created_categories += 1
            else:
                result.updated_categories += 1
            progress = True
        if not progress:
            for item in next_pending:
                result.failures.append(
                    f'分类第 {item["row_no"]} 行：找不到上级分类 slug={item["parent_slug"]}'
                )
            break
        pending_rows = next_pending
    return slug_map, path_map


def _resolve_category(level1, level2, level3, slug_map, path_map):
    parts = [part for part in (level1, level2, level3) if part]
    if not parts:
        return None
    path_key = category_path_key(parts)
    category = path_map.get(path_key)
    if category:
        return category
    slug = stable_slug(*parts)
    return slug_map.get(slug)


def _import_attributes(workbook, slug_map, path_map, result):
    if ATTRIBUTE_SHEET not in workbook.sheetnames:
        return

    attribute_sheet = workbook[ATTRIBUTE_SHEET]
    headers = [clean_text(value) for value in next(attribute_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {header: index for index, header in enumerate(headers)}
    hierarchical = '一级目录' in headers

    for row_no, row in enumerate(attribute_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if hierarchical:
            level1 = _cell(row, header_index, '一级目录')
            level2 = _cell(row, header_index, '二级目录')
            level3 = _cell(row, header_index, '三级目录')
            name = _cell(row, header_index, '属性名称')
            code = _cell(row, header_index, '属性编码')
            data_type = _cell(row, header_index, '数据类型')
            is_required = _cell(row, header_index, '必填')
            is_filterable = _cell(row, header_index, '用于筛选')
            is_list_visible = _cell(row, header_index, '列表显示')
            is_detail_visible = _cell(row, header_index, '详情显示')
            if not any((level1, level2, level3, name)):
                continue
            category = _resolve_category(level1, level2, level3, slug_map, path_map)
            if not category:
                result.failures.append(
                    f'{ATTRIBUTE_SHEET} 第 {row_no} 行：找不到分类 {" / ".join(part for part in (level1, level2, level3) if part)}'
                )
                continue
        else:
            category_slug = _cell(row, header_index, '分类slug')
            name = _cell(row, header_index, '属性名称')
            code = _cell(row, header_index, '属性编码')
            data_type = _cell(row, header_index, '数据类型')
            is_required = _cell(row, header_index, '必填')
            is_filterable = _cell(row, header_index, '用于筛选')
            is_list_visible = _cell(row, header_index, '列表显示')
            is_detail_visible = _cell(row, header_index, '详情显示')
            if not any((category_slug, name, code)):
                continue
            category = slug_map.get(category_slug) or Category.objects.filter(slug=category_slug).first()
            if not category:
                result.failures.append(f'{ATTRIBUTE_SHEET} 第 {row_no} 行：找不到分类 slug={category_slug}')
                continue

        if not name:
            result.failures.append(f'{ATTRIBUTE_SHEET} 第 {row_no} 行：属性名称不能为空')
            continue
        if not code:
            code = attribute_code(name)
        data_type_value = DATA_TYPE_BY_LABEL.get(data_type, data_type) if data_type else attribute_type(name)
        if data_type_value not in DATA_TYPE_LABELS:
            result.failures.append(f'{ATTRIBUTE_SHEET} 第 {row_no} 行：不支持的数据类型 {data_type}')
            continue
        try:
            sort_order = row_no * 10
            is_required = parse_bool(is_required, default=False)
            is_filterable = parse_bool(is_filterable, default=True)
            is_list_visible = parse_bool(is_list_visible, default=False)
            is_detail_visible = parse_bool(is_detail_visible, default=True)
            is_active = True
        except ValueError as exc:
            result.failures.append(f'{ATTRIBUTE_SHEET} 第 {row_no} 行：{exc}')
            continue

        _, created = CategoryAttribute.objects.update_or_create(
            category=category,
            code=code,
            defaults={
                'name': name,
                'data_type': data_type_value,
                'is_required': is_required,
                'is_filterable': is_filterable,
                'is_list_visible': is_list_visible,
                'is_detail_visible': is_detail_visible,
                'sort_order': sort_order,
                'is_active': is_active,
            },
        )
        if created:
            result.created_attributes += 1
        else:
            result.updated_attributes += 1


def import_categories_workbook(workbook, *, dry_run=False):
    result = CategoryImportResult()
    category_rows = _parse_category_rows(workbook, result)
    if result.failures or not category_rows:
        return result

    with transaction.atomic():
        slug_map, path_map = _upsert_categories(category_rows, result)
        _import_attributes(workbook, slug_map, path_map, result)
        if dry_run:
            transaction.set_rollback(True)

    return result


def import_categories_file(uploaded_file, *, dry_run=False):
    workbook = load_workbook(uploaded_file, data_only=True)
    return import_categories_workbook(workbook, dry_run=dry_run)
