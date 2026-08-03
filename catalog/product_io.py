from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .category_io import breadcrumb_parts, stable_slug
from .models import Category, Product, PublishStatus, SKU, StockStatus


PRODUCT_SHEET = '产品'
PRODUCT_SHEET_ALIASES = (PRODUCT_SHEET, 'Sheet1')

BASE_EXPORT_HEADERS = (
    '款式编码',
    '商品编码',
    '商品名称',
    '商品简称',
    '品牌',
    '一级目录',
    '二级目录',
    '三级目录',
    'CAS号',
    '商品状态',
    '颜色及规格',
    '颜色',
    '规格',
    '单位',
    '基本售价',
    '成本价',
    '采购价',
    '市场|吊牌价',
    '起订量',
    '销售步进',
    '库存状态',
    '库存同步',
    'SKU状态',
    '聚水潭编码',
    '店铺商品编码',
    '在线图片URL',
    '简介',
    '关键规格',
    '储存条件',
    '安全说明',
    '运输说明',
)

REQUIRED_COLUMN_GROUPS = (
    ('商品编码', 'SKU'),
    ('商品名称', '产品名称'),
)

STATUS_EXPORT_LABELS = {
    PublishStatus.PUBLISHED: '上架',
    PublishStatus.DRAFT: '草稿',
    PublishStatus.ARCHIVED: '停用',
}

STOCK_STATUS_LABELS = {choice.value: choice.label for choice in StockStatus}


@dataclass
class ProductImportResult:
    created_products: int = 0
    updated_products: int = 0
    created_skus: int = 0
    updated_skus: int = 0
    failures: list[str] = field(default_factory=list)


def clean_text(value):
    if value is None:
        return ''
    return str(value).strip()


def parse_decimal(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value).strip()).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return None


def clean_url(value):
    text = clean_text(value)
    if text.startswith(('http://', 'https://')):
        return text
    return ''


def get_any(row_values, get, *columns):
    for column in columns:
        value = clean_text(get(row_values, column))
        if value:
            return value
    return ''


def find_or_create_category(row_values, get, allow_missing):
    path = [
        clean_text(get(row_values, '一级目录')),
        clean_text(get(row_values, '二级目录')),
        clean_text(get(row_values, '三级目录')),
    ]
    category_name = next((name for name in reversed(path) if name), '')
    if not category_name:
        category_name = clean_text(get(row_values, '分类'))
    if category_name:
        candidates = list(Category.objects.filter(name=category_name, is_active=True).order_by('sort_order', 'name'))
        category = sorted(candidates, key=lambda item: item.level, reverse=True)[0] if candidates else None
        if category:
            return category
        slug = stable_slug(*[part for part in path if part]) if any(path) else stable_slug(category_name)
        return Category.objects.get_or_create(slug=slug, defaults={'name': category_name})[0]
    if allow_missing:
        return Category.objects.get_or_create(slug='uncategorized', defaults={'name': '未分类'})[0]
    return None


def product_sheet_name(workbook):
    for name in PRODUCT_SHEET_ALIASES:
        if name in workbook.sheetnames:
            return name
    return workbook.sheetnames[0] if workbook.sheetnames else None


def collect_attribute_headers(products):
    headers = []
    seen = set()
    for product in products:
        if not product.category:
            continue
        for attribute in product.category.effective_attributes():
            if attribute.name in seen:
                continue
            seen.add(attribute.name)
            headers.append(attribute.name)
    return headers


def expand_product_queryset(queryset):
    product_ids = queryset.values_list('pk', flat=True)
    return (
        Product.objects.filter(pk__in=product_ids)
        .select_related('category', 'category__parent', 'category__parent__parent')
        .prefetch_related('skus')
        .order_by('name', 'style_code')
    )


def export_products_workbook(products):
    products = list(products)
    attribute_headers = collect_attribute_headers(products)
    headers = list(BASE_EXPORT_HEADERS) + attribute_headers

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = PRODUCT_SHEET
    worksheet.append(headers)

    for product in products:
        parts = breadcrumb_parts(product.category) if product.category_id else []
        while len(parts) < 3:
            parts.append('')
        skus = list(product.skus.all().order_by('internal_sku_code'))
        if not skus:
            skus = [None]

        for sku in skus:
            row = [
                product.style_code,
                sku.internal_sku_code if sku else '',
                product.name,
                product.alias,
                product.brand,
                parts[0],
                parts[1],
                parts[2],
                product.cas_no,
                STATUS_EXPORT_LABELS.get(product.status, product.status),
                sku.sku_attribute_text if sku else '',
                sku.color if sku else '',
                sku.package_spec if sku else '',
                sku.unit if sku else '',
                sku.price if sku and sku.price is not None else '',
                sku.cost_price if sku and sku.cost_price is not None else '',
                sku.purchase_price if sku and sku.purchase_price is not None else '',
                sku.list_price if sku and sku.list_price is not None else '',
                sku.moq if sku else '',
                sku.order_step if sku else '',
                STOCK_STATUS_LABELS.get(sku.stock_status, sku.stock_status) if sku else '',
                '开启' if sku is None or sku.inventory_sync_enabled else '关闭',
                STATUS_EXPORT_LABELS.get(sku.status, sku.status) if sku else '',
                sku.jst_sku_id if sku else '',
                sku.shop_sku_id if sku else '',
                product.remote_image_url,
                product.description,
                product.spec_summary,
                product.storage_info,
                product.safety_info,
                product.shipping_info,
            ]
            if sku and attribute_headers:
                attributes = sku.attributes if isinstance(sku.attributes, dict) else {}
                attribute_map = {}
                if product.category:
                    for attribute in product.category.effective_attributes():
                        attribute_map[attribute.name] = attributes.get(attribute.code, '')
                row.extend(attribute_map.get(header, '') for header in attribute_headers)
            else:
                row.extend([''] * len(attribute_headers))
            worksheet.append(row)

    return workbook


def export_products_response(products, filename_prefix='products'):
    workbook = export_products_workbook(products)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    timestamp = timezone.localtime().strftime('%Y%m%d-%H%M%S')
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}-{timestamp}.xlsx"'
    return response


def import_products_workbook(workbook, *, dry_run=False, allow_missing_price=False, allow_missing_category=False, source_file_name=''):
    result = ProductImportResult()
    sheet_name = product_sheet_name(workbook)
    if not sheet_name:
        result.failures.append('Excel 文件中没有工作表')
        return result

    worksheet = workbook[sheet_name]
    headers = [clean_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    missing_groups = [group for group in REQUIRED_COLUMN_GROUPS if not any(column in headers for column in group)]
    if missing_groups:
        result.failures.append('缺少必填列：' + '；'.join('/'.join(group) for group in missing_groups))
        return result

    header_index = {name: idx for idx, name in enumerate(headers)}

    def get(row_values, column):
        idx = header_index.get(column)
        if idx is None or idx >= len(row_values):
            return ''
        return row_values[idx]

    with transaction.atomic():
        for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row)
            raw = {header: row_values[idx] if idx < len(row_values) else None for header, idx in header_index.items()}
            goods_code = get_any(row_values, get, '商品编码', 'SKU')
            goods_name = get_any(row_values, get, '商品名称', '产品名称')
            style_code = get_any(row_values, get, '款式编码', '产品型号', 'SKU', '商品编码') or goods_code
            category = find_or_create_category(row_values, get, allow_missing_category)
            price = parse_decimal(get(row_values, '基本售价'))
            if price is None:
                price = parse_decimal(get(row_values, '销售价格'))

            if not goods_code or not style_code or not goods_name:
                result.failures.append(f'第 {row_no} 行：SKU/商品编码、产品名称/商品名称均不能为空')
                continue
            if price is None and not allow_missing_price:
                result.failures.append(f'第 {row_no} 行：基本售价不能为空或格式不正确')
                continue
            if not category and not allow_missing_category:
                result.failures.append(f'第 {row_no} 行：分类不能为空')
                continue

            status_text = clean_text(get(row_values, '商品状态'))
            if status_text == '停用':
                status = PublishStatus.ARCHIVED
            elif status_text == '草稿':
                status = PublishStatus.DRAFT
            else:
                status = PublishStatus.PUBLISHED

            sku_status_text = clean_text(get(row_values, 'SKU状态'))
            if sku_status_text == '停用':
                sku_status = PublishStatus.ARCHIVED
            elif sku_status_text == '草稿':
                sku_status = PublishStatus.DRAFT
            else:
                sku_status = status

            stock_status_label = clean_text(get(row_values, '库存状态'))
            stock_status = next(
                (value for value, label in STOCK_STATUS_LABELS.items() if label == stock_status_label),
                StockStatus.CONFIRM,
            )

            dynamic_attributes = {}
            supplied_attribute_codes = set()
            if category:
                for attribute in category.effective_attributes():
                    if attribute.name not in header_index:
                        continue
                    supplied_attribute_codes.add(attribute.code)
                    value = clean_text(get(row_values, attribute.name))
                    if value:
                        dynamic_attributes[attribute.code] = value
                    elif attribute.is_required:
                        result.failures.append(f'第 {row_no} 行：特殊属性“{attribute.name}”不能为空')
            if result.failures and result.failures[-1].startswith(f'第 {row_no} 行：特殊属性'):
                continue

            product, product_created = Product.objects.update_or_create(
                style_code=style_code,
                defaults={
                    'name': goods_name,
                    'alias': clean_text(get(row_values, '商品简称')),
                    'cas_no': clean_text(get(row_values, 'CAS号')),
                    'category': category,
                    'brand': clean_text(get(row_values, '品牌')),
                    'remote_image_url': clean_url(get(row_values, '在线图片URL')) or clean_url(get(row_values, '图片')),
                    'image_source_url': clean_url(get(row_values, '在线图片URL')) or clean_url(get(row_values, '图片')),
                    'description': clean_text(get(row_values, '简介')),
                    'spec_summary': clean_text(get(row_values, '关键规格')) or clean_text(get(row_values, '颜色及规格')),
                    'storage_info': clean_text(get(row_values, '储存条件')),
                    'safety_info': clean_text(get(row_values, '安全说明')),
                    'shipping_info': clean_text(get(row_values, '运输说明')),
                    'source_file_name': source_file_name,
                    'source_created_by': clean_text(get(row_values, '创建人')),
                    'status': status,
                },
            )
            result.created_products += int(product_created)
            result.updated_products += int(not product_created)

            existing_sku = SKU.objects.filter(internal_sku_code=goods_code).only('attributes').first()
            merged_attributes = dict(existing_sku.attributes) if existing_sku and isinstance(existing_sku.attributes, dict) else {}
            for code in supplied_attribute_codes:
                if code in dynamic_attributes:
                    merged_attributes[code] = dynamic_attributes[code]
                else:
                    merged_attributes.pop(code, None)

            _, sku_created = SKU.objects.update_or_create(
                internal_sku_code=goods_code,
                defaults={
                    'product': product,
                    'jst_sku_id': clean_text(get(row_values, '聚水潭编码')) or goods_code,
                    'shop_sku_id': clean_text(get(row_values, '店铺商品编码')) or goods_code,
                    'source_goods_code': goods_code,
                    'source_style_code': style_code,
                    'source_goods_name': goods_name,
                    'sku_attribute_text': get_any(row_values, get, '颜色及规格', '产品规格'),
                    'color': clean_text(get(row_values, '颜色')),
                    'package_spec': get_any(row_values, get, '规格', '产品规格'),
                    'unit': clean_text(get(row_values, '单位')),
                    'price': price,
                    'cost_price': parse_decimal(get(row_values, '成本价')),
                    'purchase_price': parse_decimal(get(row_values, '采购价')),
                    'list_price': parse_decimal(get(row_values, '市场|吊牌价')),
                    'moq': parse_decimal(get(row_values, '起订量')) or parse_decimal(get(row_values, '建议采购数')) or Decimal('1'),
                    'order_step': parse_decimal(get(row_values, '销售步进')) or Decimal('1'),
                    'stock_status': stock_status,
                    'inventory_sync_enabled': clean_text(get(row_values, '库存同步')) != '关闭',
                    'attributes': merged_attributes,
                    'source_raw_row': {key: clean_text(value) for key, value in raw.items()},
                    'status': sku_status,
                },
            )
            result.created_skus += int(sku_created)
            result.updated_skus += int(not sku_created)

        if dry_run:
            transaction.set_rollback(True)

    return result


def import_products_file(uploaded_file, *, dry_run=False, allow_missing_price=False, allow_missing_category=False):
    workbook = load_workbook(uploaded_file, data_only=True)
    source_file_name = getattr(uploaded_file, 'name', '') or ''
    return import_products_workbook(
        workbook,
        dry_run=dry_run,
        allow_missing_price=allow_missing_price,
        allow_missing_category=allow_missing_category,
        source_file_name=source_file_name,
    )
